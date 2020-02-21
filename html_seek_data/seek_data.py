from selenium import webdriver
import time
from bs4 import BeautifulSoup
import requests
from lxml import etree
import re
from openpyxl import load_workbook



def get_url_page(num):
    """
    获取第几页的数据
    :param page:
    :return:
    """

    # url_list = []
    page = num*2 - 1
    s = 1 + 60*(num-1)
    url = "https://search.jd.com/Search?keyword=%E8%BD%A6%E5%8E%98%E5%AD%90&enc=utf-8&qrst=1&rt=1&stop=1&vt=2&stock=1&page="+str(page)+"&s="+str(s)+"&click=0"
    # url = requests.utils.unquote(url)

    return url
#

header = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36",
                  "content-type": "application/x-www-form-urlencoded", "accept-language": "zh-CN,zh;q=0.9"
                  }


def get_item_url():
    """
    根据搜索的商品，把页面每个商品的数据都获取下来
    :return:
    """
    detail_url = ['https://search.jd.com/Search?keyword=%E8%BD%A6%E5%8E%98%E5%AD%90&enc=utf-8&wq=%E8%BD%A6%E5%8E%98%E5%AD%90&pvid=a4aa735b48324cb1b77ed694579e3ebd']
    for url_1 in detail_url:

        header = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36",
                  "content-type": "application/x-www-form-urlencoded", "accept-language": "zh-CN,zh;q=0.9"
                  }
        data = {"appname":"jdwebm_pv","jdkey":"","whwswswws":"zP/PeIYrAnGbFS9MxseFvhw==","businness":"pcSearch","body":{"sid":"96509f4176d09b4a3f05c6e0d463bd48","squence":"3","create_time":"1580871351126","shshshfpa":"98940fa8-2fca-21a3-1a10-1684039dd35c-1580871336","ecflag":"n","whwswswws":"zP/PeIYrAnGbFS9MxseFvhw==","browser_info":"0dcc2d612bbed3970c0e3800135ec2d4","page_name":"https://search.jd.com/Search","page_param":"keyword=车厘子&enc=utf-8&pvid=d65169487c7948698c54ba5a3af65c69","cookie_pin":"","msdk_version":"2.3.6","wid":"","pv_referer":"https://search.jd.com/Search?keyword=%25E8%25BD%25A6%25E5%258E%2598%25E5%25AD%2590&enc=utf-8&suggest=1.def.0.V15--12s0,20s0,38s0,97s0&wq=chelizi&pvid=db7a550aa31040f6a13f5b2a177dba7d"}}
        res = requests.post(url_1, headers = header)
        res.encoding = "gb2312"

        soup = BeautifulSoup(res.text, "html.parser")

        all = soup.find_all("a")  # 读取商品页的数据

        #非法链接
        invalidLink1 = '#'
        invalidLink2 = 'javascript:void(0)'
        count = 0
        item_link = []
        for i in all:
            # count = 0
            if i.find("href") != -1 and i.find("href")!= invalidLink1 and i.find("href")!=invalidLink2:
                link = i.get("href")
                if link.find("item") !=-1 and link.find("comment") ==-1:
                    link = "https://"+link[2::]
                    item_link.append(link) # 把商品链接取出来添加到空列表中

    item_link = set(item_link)
    return item_link




def get_item_data(url):
    """
    获取物品详细信息
    :param url: 商品链接
    :return:
    """
    driver = webdriver.Chrome()
    driver.minimize_window()
    driver.get(url)
    time.sleep(3)
    try:
        name = driver.find_element_by_class_name("sku-name").text
        # print(name)
        price = driver.find_element_by_class_name("p-price").text
        # print(price)
        plus_price = driver.find_element_by_class_name("p-price-plus").text
        if len(plus_price) ==0:
            plus_price = '无超级会员价格'

        # print(plus_price)
        store_status = driver.find_element_by_class_name("store-prompt").text
        # print(store_status)
        return name, price, plus_price, store_status
    except Exception as e:
        raise e


def write_excel(row, col, values):
    """
    往excel 指定位置写入值
    :param row: 行
    :param col: 列
    :param value: 需要写入的值
    :return:
    """
    wb = load_workbook("../jd_item.xlsx")
    sheet = wb["jd"]
    sheet.cell(row, col).value = values
    wb.save("jd_item.xlsx")
    wb.close()




if __name__ == '__main__':
    item_url = get_item_url()
    # item_url = ['https://item.jd.com/100002662266.html', 'https://item.jd.com/100001864407.html', 'https://item.jd.com/5522605.html']

    i = len(item_url)


    num = 2
    for url in item_url:
        item_data = get_item_data(url)
        print(url)
        name = item_data[0]
        price = item_data[1]
        plus_price = item_data[2]
        status = item_data[3]
        write_excel(num, 1,name)
        write_excel(num, 2, price)
        write_excel(num, 3, plus_price)
        write_excel(num, 4, status)
        write_excel(num, 5, url)
        num = num + 1



