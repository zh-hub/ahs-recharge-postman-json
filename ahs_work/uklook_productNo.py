from openpyxl import load_workbook
import requests
import json

wb = load_workbook("test.xlsx")
sheet = wb["test"]
col = 1
row = 1
n = 0
# product_no = ["20191220164135532543"]
product_no=[]
while n<=40:
      product_no_1 = sheet.cell(row+n, col).value
      n = n + 1
      if product_no_1 != None:
            product_no.append(product_no_1)
#
# print(len(product_no))
url = "http://47.96.53.33:8080/inventoryservice/inventory/unlock"   # uat
# url = "http://120.55.138.130:8080/inventoryservice/inventory/unlock"  # fat
data = {"operatorId":0,"inventorySerialNumbers":product_no}
data = json.dumps(data)
res = requests.post(url,data=data)
print(res.text)